
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags = CREATE_NO_WINDOW   
driver = webdriver.Chrome(service = serv)


import time, datetime, os
import openpyxl   


now = str(datetime.datetime.now())[:16]
folderName = now.replace(":", "_")
os.mkdir(folderName)


key_word = ["2024년"]

wb = openpyxl.Workbook()  #워크북

for i in range(len(key_word)):
    ws = wb.create_sheet()   #워크시트 
    ws.title = key_word[i]
    ws.column_dimensions["A"].width = 90   #A열 너비 조절
    
    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i]
    driver.get(url)
    time.sleep(2)
                  
    list_news = driver.find_element("class name", "list_news")
    news_boxes = list_news.find_elements("class name", "bx")
    
    for j in range(len(news_boxes)):      
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])   #찾는 개체까지 스크롤 다운
        file = f"{folderName}/{i+1}_{key_word[i]}-{j+1}.png"  #이미지 파일 이름 미리 지정
        news_boxes[j].screenshot(file)   #스크린샷해서 저장

        ws.row_dimensions[j+1].height = 100   #행 높이 조절
        img = openpyxl.drawing.image.Image(file)   #스크린샷을 엑셀 삽입 가능한 이미지로 변환
        ws.add_image(img, f'A{j+1}')   #A열 해당 셀(j+1)에 이미지 삽입


        title = news_boxes[j].find_element("class name", "news_tit")   #링크 삽입을 위한 news_tit 개체를 title로 저장
        print(j+1, title.text)
        
        link = title.get_attribute("href")   #title 개체의 링크를 저장
        ws[f'B{j+1}'].value = "기사링크"   
        ws[f'B{j+1}'].hyperlink = link   #B열 해당 셀(j+1)에 링크 삽입


    print()
    
wb.remove(wb["Sheet"])   #필요없는 첫번째 시트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx")   #엑셀 저장
